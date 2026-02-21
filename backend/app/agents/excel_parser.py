"""
AI-powered Excel parsing agent using Google Gemini.
LLM handles: header detection, fuzzy column→parameter mapping, asset detection.
Deterministic code handles: value parsing, validation, file I/O.
"""
import json
import logging
import os
import re
from io import BytesIO

import google.generativeai as genai
import openpyxl
from openpyxl.worksheet.worksheet import Worksheet

from app.models.schemas import (
    ColumnMapping,
    LLMMappingResponse,
    ParsedCell,
    ParseResponse,
    UnmappedColumn,
)
from app.utils.registry import load_assets, load_parameters
from app.utils.value_parser import parse_value, validate_value

logger = logging.getLogger(__name__)

# Configure Gemini
genai.configure(api_key=os.environ["GEMINI_API_KEY"])
MODEL = os.getenv("GEMINI_MODEL", "gemini-1.5-flash")


def _build_mapping_prompt(headers: list[str], sheet_name: str = "Sheet1") -> str:
    parameters = load_parameters()
    assets = load_assets()

    param_summary = json.dumps(
        [{"name": p["name"], "display_name": p["display_name"], "unit": p["unit"]} for p in parameters],
        indent=2,
    )
    asset_summary = json.dumps(
        [{"name": a["name"], "display_name": a["display_name"], "type": a["type"]} for a in assets],
        indent=2,
    )

    return f"""You are an expert industrial data analyst for an ESG platform called LatSpace.

Your job is to map Excel column headers from a factory data spreadsheet to our canonical parameter registry.

## Parameter Registry
{param_summary}

## Asset Registry
{asset_summary}

## Task
The Excel sheet "{sheet_name}" has these column headers (with their column index):
{json.dumps({str(i): h for i, h in enumerate(headers)}, indent=2)}

For EACH header, determine:
1. Which parameter from the registry it maps to (or null if unmapped)
2. Which asset it refers to (or null if none/plant-level)
3. Confidence level: "high" (exact/near-exact), "medium" (reasonable guess), "low" (unclear)
4. Brief reasoning

Rules:
- Be aggressive about fuzzy matching: "Coal Used (MT)" → coal_consumption, "COAL CONSMPTN" → coal_consumption
- Detect embedded asset names: "Coal Consumption AFBC-1" → param=coal_consumption, asset=AFBC-1
- "Steam (Boiler 2)" → param=steam_generation, asset=AFBC-2 (Boiler 2 = AFBC-2)
- "Power TG1" → param=power_generation, asset=TG-1
- Generic columns like "Date", "Day", "Comments", "Sr No" → unmapped (null param)
- Return the header_row_index (0-based index of the row that contains headers, usually 0 unless there are title rows)

## Output Format
Return ONLY valid JSON matching this exact schema. No markdown, no explanation:
{{
  "header_row_index": 0,
  "mappings": [
    {{
      "col_index": 0,
      "original_header": "Coal Consumption AFBC-1",
      "param_name": "coal_consumption",
      "asset_name": "AFBC-1",
      "confidence": "high",
      "reasoning": "Exact parameter name with explicit asset suffix"
    }}
  ],
  "unmapped_headers": ["Comments", "Sr No"],
  "notes": "Any overall observations about the file structure"
}}"""


def _extract_headers_and_data(
    ws: Worksheet, max_scan_rows: int = 10
) -> tuple[int, list[str], list[list]]:
    """
    Scan first max_scan_rows to find the header row.
    Returns (header_row_idx, headers, all_data_rows).
    Uses a heuristic: the header row has the most non-empty string cells.
    """
    all_rows = list(ws.iter_rows(values_only=True))
    if not all_rows:
        return 0, [], []

    best_row_idx = 0
    best_score = -1

    for i, row in enumerate(all_rows[:max_scan_rows]):
        score = sum(1 for cell in row if cell is not None and isinstance(cell, str) and cell.strip())
        if score > best_score:
            best_score = score
            best_row_idx = i

    headers = [str(c).strip() if c is not None else "" for c in all_rows[best_row_idx]]
    data_rows = all_rows[best_row_idx + 1 :]
    return best_row_idx, headers, data_rows


def _call_gemini_for_mapping(headers: list[str], sheet_name: str) -> LLMMappingResponse:
    """Single LLM call to map all headers at once."""
    prompt = _build_mapping_prompt(headers, sheet_name)
    model = genai.GenerativeModel(MODEL)

    try:
        response = model.generate_content(
            prompt,
            generation_config=genai.types.GenerationConfig(
                temperature=0.1,  # low temperature for deterministic mapping
                response_mime_type="application/json",
            ),
        )
        raw = response.text.strip()
        # Strip markdown code fences if present
        raw = re.sub(r"^```(?:json)?\s*", "", raw)
        raw = re.sub(r"\s*```$", "", raw)
        data = json.loads(raw)
        return LLMMappingResponse(**data)
    except Exception as e:
        logger.error(f"Gemini mapping call failed: {e}")
        # Fallback: return empty mappings
        return LLMMappingResponse(
            header_row_index=0,
            mappings=[
                ColumnMapping(col_index=i, original_header=h, confidence="low", reasoning="LLM unavailable")
                for i, h in enumerate(headers)
            ],
        )


def parse_excel(file_bytes: bytes, filename: str) -> ParseResponse:
    """
    Main entry point: parse an Excel file and return structured data.
    Supports multi-sheet workbooks.
    """
    wb = openpyxl.load_workbook(BytesIO(file_bytes), data_only=True)
    all_parsed: list[ParsedCell] = []
    all_unmapped: list[UnmappedColumn] = []
    all_warnings: list[str] = []
    all_duplicates: list[str] = []
    final_header_row = 0
    seen_param_asset: dict[str, int] = {}  # duplicate detection

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        header_row_idx, headers, data_rows = _extract_headers_and_data(ws)

        if not headers or all(h == "" for h in headers):
            all_warnings.append(f"Sheet '{sheet_name}': No headers found, skipped.")
            continue

        if header_row_idx > 0:
            all_warnings.append(
                f"Sheet '{sheet_name}': Rows 0–{header_row_idx - 1} appear to be title/metadata rows, skipped."
            )

        # One LLM call per sheet
        mapping_result = _call_gemini_for_mapping(headers, sheet_name)
        final_header_row = mapping_result.header_row_index

        # Build lookup: col_index → ColumnMapping
        col_map: dict[int, ColumnMapping] = {m.col_index: m for m in mapping_result.mappings}

        # Track unmapped columns
        for mapping in mapping_result.mappings:
            if mapping.param_name is None:
                all_unmapped.append(
                    UnmappedColumn(
                        col=mapping.col_index,
                        header=mapping.original_header,
                        reason=mapping.reasoning or "No matching parameter found",
                    )
                )

        # Duplicate detection
        for mapping in mapping_result.mappings:
            if mapping.param_name:
                key = f"{mapping.param_name}::{mapping.asset_name or 'plant'}"
                if key in seen_param_asset:
                    all_duplicates.append(
                        f"Sheet '{sheet_name}': Duplicate param+asset '{key}' "
                        f"at col {mapping.col_index} and col {seen_param_asset[key]}"
                    )
                else:
                    seen_param_asset[key] = mapping.col_index

        # Parse data rows
        for row_offset, row in enumerate(data_rows):
            actual_row_num = header_row_idx + 1 + row_offset + 2  # 1-indexed for display

            for col_idx, cell_value in enumerate(row):
                mapping = col_map.get(col_idx)
                if mapping is None or mapping.param_name is None:
                    continue

                parsed = parse_value(cell_value)
                raw_str = str(cell_value) if cell_value is not None else ""

                # Validation warnings
                val_warnings = validate_value(mapping.param_name, parsed)
                for w in val_warnings:
                    all_warnings.append(f"Row {actual_row_num}, col {col_idx}: {w}")

                all_parsed.append(
                    ParsedCell(
                        row=actual_row_num,
                        col=col_idx,
                        param_name=mapping.param_name,
                        asset_name=mapping.asset_name,
                        raw_value=raw_str,
                        parsed_value=parsed,
                        confidence=mapping.confidence,
                    )
                )

    return ParseResponse(
        status="success",
        header_row=final_header_row,
        parsed_data=all_parsed,
        unmapped_columns=all_unmapped,
        warnings=all_warnings,
        duplicate_flags=all_duplicates,
    )