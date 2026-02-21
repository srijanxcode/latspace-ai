from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill

OUTPUT_DIR = Path(__file__).parent / "test_data"
OUTPUT_DIR.mkdir(exist_ok=True)

def create_clean_data():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "January 2024"
    headers = ["Date", "Coal Consumption", "Coal GCV", "Steam Generation",
               "Power Generation", "CO2 Emissions", "Boiler Efficiency", "Water Consumption"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="4472C4")
    data = [
        ["2024-01-01", 120.5, 3850, 450.2, 85.3, 312.1, 82.5, 1200],
        ["2024-01-02", 118.0, 3820, 445.0, 84.1, 305.8, 81.9, 1180],
        ["2024-01-03", 125.3, 3900, 460.5, 87.2, 325.4, 83.1, 1250],
        ["2024-01-04", "N/A", 3780, 438.0, 82.5, "N/A", 80.5, 1150],
        ["2024-01-05", 119.8, 3810, "N/A", 83.7, 310.2, 82.0, 1190],
    ]
    for row in data:
        ws.append(row)
    wb.save(OUTPUT_DIR / "clean_data.xlsx")
    print("✅ clean_data.xlsx created")

def create_messy_data():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Monthly Report"
    ws.append(["ACME INDUSTRIES — MONTHLY OPERATIONAL REPORT"])
    ws.append(["Period: January 2024", "", "Unit: Metric Tonnes / MWh"])
    ws.append([])
    messy_headers = ["Sl No", "COAL CONSMPTN", "Coal GCV (kcal)", "Stm Gen T/hr",
                     "Pwr Gen (MWh)", "CO2 Emssns", "Effcncy %", "H2O Consmptn KL", "Remarks"]
    ws.append(messy_headers)
    data = [
        [1, "1,234.56", "3,850", "450.2", "85.3", "312.1", "82.5%", "1200", "Normal ops"],
        [2, "1180",     "3820",  "445",   "84.1", "305.8", "81.9%", "1,180", ""],
        [3, "1253",     "3,900", "460.5", "87.2", "325.4", "83.1%", "1250", "Peak load"],
        [4, "-",        "3780",  "438",   "82.5", "-",     "80.5%", "1150", "Data missing"],
        [5, "YES",      "3810",  "N/A",   "83.7", "310.2", "82%",  "1190", ""],
    ]
    for row in data:
        ws.append(row)
    wb.save(OUTPUT_DIR / "messy_data.xlsx")
    print("✅ messy_data.xlsx created")

def create_multi_asset_data():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Assets Combined"
    headers = [
        "Date",
        "Coal Consumption AFBC-1", "Coal Consumption AFBC-2",
        "Steam Generation AFBC-1", "Steam (Boiler 2)",
        "Power TG1",               "Power Generation TG-2",
        "CO2 Emissions AFBC-1",    "CO2 Emissions AFBC-2",
        "Efficiency Boiler 1",     "Efficiency AFBC-2",
        "Comments"
    ]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    data = [
        ["2024-01-01", 65.2, 55.3, 240.1, 210.5, 45.2, 40.1, 169.5, 143.8, 83.2, 81.5, "Normal"],
        ["2024-01-02", 63.0, 55.0, 238.0, 207.0, 44.1, 40.0, 163.8, 142.0, 82.8, 81.2, ""],
        ["2024-01-03", "N/A", 58.1, "N/A", 220.3, 46.5, 40.7, "N/A", 151.1, "N/A", 82.0, "B1 offline"],
        ["2024-01-04", 67.5, 57.8, 244.2, 218.6, 47.0, 41.5, 175.5, 150.3, 83.8, 82.1, ""],
    ]
    for row in data:
        ws.append(row)
    wb.save(OUTPUT_DIR / "multi_asset.xlsx")
    print("✅ multi_asset.xlsx created")

if __name__ == "__main__":
    print("Creating test Excel files...")
    create_clean_data()
    create_messy_data()
    create_multi_asset_data()
    print("All files saved to:", OUTPUT_DIR)
